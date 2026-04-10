import argparse
import json
import os
import shutil
import sys
import webbrowser
from datetime import date
from urllib.parse import quote
from copy import copy
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

MAINTAINER_EMAIL = "your.email@example.com"


def report_error(error_msg):
    print(f"\n[ERROR] {error_msg}")
    choice = input(
        "\nAn error occurred during test generation. Would you like to report this to the maintainer? (y/N): "
    )
    if choice.lower() == "y":
        subject = quote("Brycen TestGen Script Error")
        body = quote(f"Error: {error_msg}\n\nPython: {sys.version}\nOS: {sys.platform}")
        webbrowser.open(f"mailto:{MAINTAINER_EMAIL}?subject={subject}&body={body}")


def generate_tests(input_json, template_path, output_path):
    # JSON data validation
    try:
        with open(input_json, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        raise Exception(f"Error reading JSON: {e}")

    if "tests" not in data:
        raise Exception("Invalid JSON format: missing 'tests' key.")

    meta = data.get("meta", {})
    system_name = meta.get("systemName", "")
    sub_system_name = meta.get("subSystemName", "")
    author = meta.get("author", "")
    current_date = date.today().strftime("%Y/%m/%d")

    # Filter tests based on the "selected" boolean (default to True if missing)
    all_tests = data["tests"]
    selected_tests = [t for t in all_tests if t.get("selected", True)]

    if not selected_tests:
        print(
            "No tests selected (all 'selected' fields are false or no tests provided). Aborting."
        )
        return

    data["tests"] = selected_tests

    # Template copying logic
    if not os.path.exists(template_path):
        raise Exception(f"Template not found: {template_path}")

    shutil.copyfile(template_path, output_path)
    print(f"Created new file from template: {output_path}")

    # Load the new workbook
    try:
        wb = load_workbook(output_path)
        ws = wb.active
    except Exception as e:
        raise Exception(f"Error loading workbook: {e}")

    # Column Mapping based on template analysis (Row 6 headers)
    COL_NO = 1
    COL_CLASS = 3
    COL_CONTENT = 8
    COL_METHOD = 13
    COL_RESULT = 24

    # Data starts at row 7 as requested
    start_row = 7
    # Reference row for styles (Row 11 has good formatting examples)
    ref_row = 11

    # Pre-load styles from reference row for each target column
    styles = {}
    for col in [COL_NO, COL_CLASS, COL_CONTENT, COL_METHOD, COL_RESULT]:
        ref_cell = ws.cell(row=ref_row, column=col)
        styles[col] = {
            "alignment": copy(ref_cell.alignment),
            "font": copy(ref_cell.font),
            "border": copy(ref_cell.border),
            "fill": copy(ref_cell.fill),
            "number_format": ref_cell.number_format,
        }

    for i, test in enumerate(data["tests"], start=1):
        current_row = start_row + i - 1

        mapping = {
            COL_NO: i,
            COL_CLASS: test.get("Classification", ""),
            COL_CONTENT: test.get("Content", ""),
            COL_METHOD: test.get("Test method", ""),
            COL_RESULT: test.get("Test result", ""),
        }

        for col, value in mapping.items():
            cell = ws.cell(row=current_row, column=col, value=value)
            # Apply template styles
            if col in styles:
                cell.alignment = styles[col]["alignment"]
                cell.font = styles[col]["font"]
                cell.border = styles[col]["border"]
                cell.fill = styles[col]["fill"]
                cell.number_format = styles[col]["number_format"]

    # Re-inject Data Validation (dropdowns) that openpyxl stripped
    try:
        last_row = start_row + len(data["tests"]) - 1

        # 1. Tester Dropdown (Column AI = 35) referencing List sheet Column B
        dv_tester = DataValidation(
            type="list", formula1="List!$B$1:$B$20", allow_blank=True
        )
        ws.add_data_validation(dv_tester)
        dv_tester.add(f"AI{start_row}:AI{last_row}")

        # 2. Result Dropdown (Column AM = 39) referencing List sheet Column A
        dv_result = DataValidation(
            type="list", formula1="List!$A$1:$A$10", allow_blank=True
        )
        ws.add_data_validation(dv_result)
        dv_result.add(f"AM{start_row}:AM{last_row}")

    except Exception as e:
        print(f"Warning: Could not re-inject dropdowns: {e}")

    if system_name or author or current_date:
        try:
            ws_title = wb["Title"]
            ws_title["I13"] = system_name
            ws_title["L18"] = current_date
            ws_title["L19"] = current_date
            ws_title["P18"] = author
            ws_title["P19"] = author
        except Exception as e:
            print(f"Warning: Could not write metadata to Title sheet: {e}")

        try:
            ws_testcase = wb["TestCase"]
            ws_testcase["Q1"] = system_name
            ws_testcase["Q2"] = sub_system_name
            ws_testcase["AI1"] = current_date
            ws_testcase["AP1"] = author
        except Exception as e:
            print(f"Warning: Could not write metadata to TestCase sheet: {e}")

        try:
            ws_ng_report = wb["NG Report"]
            ws_ng_report["Q1"] = system_name
            ws_ng_report["Q2"] = sub_system_name
            ws_ng_report["AI1"] = current_date
            ws_ng_report["AP1"] = author
        except Exception as e:
            print(f"Warning: Could not write metadata to NG Report sheet: {e}")

    try:
        wb.save(output_path)
    except Exception as e:
        raise Exception(f"Error saving workbook: {e}")

    print(f"Successfully wrote {len(data['tests'])} test cases to {output_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate Excel test cases from JSON.")
    parser.add_argument("--input", required=True, help="Path to input JSON file")
    parser.add_argument("--template", required=True, help="Path to Excel template file")
    parser.add_argument("--output", required=True, help="Path to output Excel file")

    args = parser.parse_args()
    try:
        generate_tests(args.input, args.template, args.output)
    except Exception as e:
        report_error(str(e))
        sys.exit(1)
